from openpyxl import load_workbook, Workbook



def run_tests_from_excel(*args, **kwargs):
    try:
        _run_tests_from_excel(*args, **kwargs).send(None)
    except StopIteration:
        pass



def params_from_excel(infilename):
    wb = load_workbook(infilename, read_only=True, data_only=True)

    op_params = Opparams(*(i.value for i in wb["operational_params"][2]))
    test_param_list = []
    test_param_dict = {}
    for count, row in enumerate(wb["params"]):
        if row[0].value not in actions:
            continue
        else:
            tparams = Testparams(count + 1, *(i.value for i in row))
            if tparams.rbw_Hz is None and tparams.span_Hz is not None:
                tparams=tparams._replace(rbw_Hz=max(tparams.span_Hz/100,100))
            if tparams.vbw_Hz is None:
                tparams = tparams._replace(vbw_Hz=tparams.rbw_Hz)
            test_param_list.append(tparams)
            test_param_dict[tparams.test_num] = tparams[:-3]

    wb.close()
    return test_param_list,test_param_dict,op_params



#from openpyxl import Workbook
wb = load_workbook(r'C:\_KM_root_lap\Docs_Personal\sampleExcel.xlsx')

# grab the active worksheet
ws = wb.active

ws = wb['secondSheet']
ws['H6'] = "sheet2222 New text right here"

ws = wb['firstSheet']
ws['G4'] = "first sheeet 11111"

# Data can be assigned directly to cells
ws['A6'] = "New text right here"
#"NEW TEXT RIGHT HERE"

# Rows can also be appended
ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save(r"C:\_KM_root_lap\Docs_Personal\sampleExcel.xlsx")